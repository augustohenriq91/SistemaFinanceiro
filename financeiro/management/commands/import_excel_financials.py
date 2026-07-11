import datetime
import os
import re
import shutil
import unicodedata
from decimal import Decimal
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
from financeiro.models import Categoria, Conta, Receita, Despesa, EmprestimoCartao, ParcelaEmprestimo

try:
    import openpyxl
except ImportError:
    openpyxl = None


class Command(BaseCommand):
    help = 'Limpa o banco, faz backup do SQLite e importa dados da planilha Planilha_mensal.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            help='Caminho para a planilha Excel. Padrão: Planilha_mensal.xlsx',
            default='Planilha_mensal.xlsx',
        )
        parser.add_argument(
            '--sheet',
            help='Nome da aba com os lançamentos. Padrão: Janeiro',
            default='Janeiro',
        )
        parser.add_argument(
            '--all',
            action='store_true',
            help='Importa todas as abas válidas da planilha Excel',
        )
        parser.add_argument(
            '--noinput',
            action='store_true',
            help='Não perguntar antes de limpar o banco',
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            self.stderr.write('openpyxl não está instalado. Instale com pip install openpyxl')
            return

        file_path = options['file']
        sheet_name = options['sheet']
        all_sheets = options['all']

        self.stdout.write(f'Importando planilha: {file_path} ' + ('(todas as abas)' if all_sheets else f'(aba: {sheet_name})'))

        if not options['noinput']:
            confirm = input('Isto irá limpar Receitas, Despesas, Contas e Categorias. Continuar? [s/N]: ')
            if confirm.lower() not in ('s', 'sim'):
                self.stdout.write('Operação cancelada.')
                return

        self._backup_database()

        wb = openpyxl.load_workbook(file_path, data_only=True)
        with transaction.atomic():
            self._clear_data()
            self._ensure_default_categories_and_accounts()

            if all_sheets:
                processed = self._process_all_sheets(wb)
                if processed == 0:
                    self.stderr.write('Nenhuma aba válida foi importada.')
                    return
            else:
                if sheet_name not in wb.sheetnames:
                    self.stderr.write(f'Aba não encontrada: {sheet_name}')
                    return
                ws = wb[sheet_name]
                if not self._process_sheet(ws, sheet_name):
                    self.stderr.write(f'Aba {sheet_name} não pôde ser importada.')
                    return

        self.stdout.write(self.style.SUCCESS('Importação concluída.'))

    def _process_all_sheets(self, wb):
        processed_count = 0
        for sheet_name in wb.sheetnames:
            if self._should_skip_sheet(sheet_name):
                self.stdout.write(f'Skipping sheet: {sheet_name}')
                continue
            ws = wb[sheet_name]
            self.stdout.write(f'Importando aba: {sheet_name}')
            if self._process_sheet(ws, sheet_name):
                processed_count += 1
            else:
                self.stdout.write(f'Aba ignorada (sem cabeçalho válido): {sheet_name}')
        return processed_count

    def _should_skip_sheet(self, sheet_name):
        normalized = self._normalize_text(sheet_name)
        return normalized == 'FATURAMENTO' or normalized == 'JANEIRO2'

    def _backup_database(self):
        db_path = settings.BASE_DIR / 'db.sqlite3'
        if db_path.exists():
            backup_name = f'db_backup_{datetime.datetime.now():%Y%m%d_%H%M%S}.sqlite3'
            backup_path = settings.BASE_DIR / backup_name
            shutil.copy2(db_path, backup_path)
            self.stdout.write(f'Backup do banco criado em {backup_name}')
        else:
            self.stdout.write('Arquivo de banco não encontrado, pulando backup.')

    def _clear_data(self):
        self.stdout.write('Limpando dados existentes...')
        Receita.objects.all().delete()
        Despesa.objects.all().delete()
        EmprestimoCartao.objects.all().delete()
        ParcelaEmprestimo.objects.all().delete()
        Conta.objects.all().delete()
        Categoria.objects.all().delete()

    def _ensure_default_categories_and_accounts(self):
        self.stdout.write('Criando contas e categorias padrão...')
        self.contas = {
            'Mercado Pago': Conta.objects.create(nome='Mercado Pago', tipo='banco', saldo_inicial=0),
            'NuBank': Conta.objects.create(nome='NuBank', tipo='banco', saldo_inicial=0),
            'Inter': Conta.objects.create(nome='Inter', tipo='banco', saldo_inicial=0),
            'Santander': Conta.objects.create(nome='Santander', tipo='banco', saldo_inicial=0),
            'Carteira': Conta.objects.create(nome='Carteira', tipo='carteira', saldo_inicial=0),
        }

        self.categorias = {
            'Receitas': Categoria.objects.create(nome='Receitas', tipo='receita'),
            'Despesas': Categoria.objects.create(nome='Despesas', tipo='despesa'),
            'Cartões': Categoria.objects.create(nome='Cartões', tipo='despesa'),
            'Investimentos': Categoria.objects.create(nome='Investimentos', tipo='despesa'),
            'Emprestimos': Categoria.objects.create(nome='Emprestimos', tipo='despesa'),
        }

    def _process_sheet(self, ws, sheet_name=None):
        self.stdout.write(f'Processando aba: {sheet_name or ws.title}')
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 16:
            self.stderr.write('A planilha parece não ter linhas suficientes para importar.')
            return False

        header_line = None
        for i, row in enumerate(rows):
            if row and len(row) > 1 and row[1] and self._normalize_text(row[1]) == 'DESCRICAO':
                header_line = i
                break

        if header_line is None:
            return False

        self.stdout.write(f'Cabeçalho encontrado na linha {header_line + 1}')
        for row_index, row in enumerate(rows[header_line + 1:], start=header_line + 2):
            if not row or all(cell is None for cell in row):
                continue

            self._process_despesa_row(row)
            self._process_receita_row(row)
            self._process_extra_row(row)

        return True

    def _process_despesa_row(self, row):
        if len(row) < 5:
            return

        descricao = row[1]
        tipo = row[2]
        valor = row[3]
        pago = row[4]

        if descricao and tipo and valor is not None:
            data = self._find_date_in_row(row)
            if data is None:
                self.stderr.write(f'Ignorando despesa sem data: {descricao}')
                return

            categoria = self._get_or_create_categoria(str(tipo).strip(), 'despesa')
            conta = self._get_or_create_conta_from_row(row)
            self._create_despesa(descricao, valor, data, categoria, conta, bool(pago))

    def _process_receita_row(self, row):
        if len(row) < 11:
            return

        descricao = row[6]
        pagador = row[7]
        tipo = row[8]
        valor = row[9]
        recebido = row[10]

        if descricao and tipo and valor is not None:
            data = self._find_date_in_row(row)
            if data is None:
                self.stderr.write(f'Ignorando receita sem data: {descricao}')
                return

            categoria = self._get_or_create_categoria(str(tipo).strip(), 'receita')
            conta = self._get_or_create_conta_from_row(row)
            self._create_receita(descricao, valor, data, categoria, conta, bool(recebido))

    def _process_extra_row(self, row):
        if len(row) < 18:
            return

        descricao = row[13]
        if not descricao or not isinstance(descricao, str):
            return

        descricao_norm = self._normalize_text(descricao)
        if descricao_norm.startswith('DESCRICAO') or descricao_norm in ('VALOR', 'DATA', 'PAGADOR', 'BANCO'):
            return

        valor = row[14]
        data = row[15]
        pagador = row[16]
        banco = row[17]

        if valor is None:
            return

        data = self._normalize_date(data) if data is not None else None
        if data is None:
            data = self._find_date_in_row(row)

        if data is None:
            self.stderr.write(f'Linha extra ignorada sem data válida: {descricao}')
            return

        if 'EMPREST' in descricao_norm:
            self._create_emprestimo(descricao, valor, data, pagador, banco, row)
        else:
            categoria = self._get_or_create_categoria('Receitas', 'receita')
            conta = self._get_or_create_conta_from_row(row)
            self._create_receita(descricao, valor, data, categoria, conta, True)

    def _create_emprestimo(self, descricao, valor, data, pessoa, banco, row):
        pessoa_text = str(pessoa).strip() if pessoa else 'Importado'
        banco_text = str(banco).strip() if banco else None
        cartao_utilizado = self.contas.get(banco_text) if banco_text else None
        if not cartao_utilizado:
            cartao_utilizado = self.contas.get('Carteira')

        conta_recebimento = cartao_utilizado or self.contas['Carteira']
        match = re.search(r'(\d{1,2})\s*/\s*(\d{1,2})', descricao)
        numero_parcela = int(match.group(1)) if match else 1
        quantidade_parcelas = int(match.group(2)) if match else 1
        valor_decimal = Decimal(str(valor))
        valor_total = (valor_decimal * Decimal(quantidade_parcelas)).quantize(Decimal('0.01')) if quantidade_parcelas > 1 else valor_decimal
        data_compra = self._normalize_date(data)
        dia_vencimento_cartao = getattr(cartao_utilizado, 'dia_vencimento_fatura', None) or 10

        emprestimo = EmprestimoCartao.objects.create(
            pessoa=pessoa_text,
            banco=banco_text or (cartao_utilizado.nome if cartao_utilizado else 'Importado'),
            cartao_utilizado=cartao_utilizado,
            conta_recebimento=conta_recebimento,
            descricao=str(descricao).strip(),
            valor_total=valor_total,
            quantidade_parcelas=quantidade_parcelas,
            data_compra=data_compra,
            dia_vencimento_cartao=dia_vencimento_cartao,
            observacao='Importado da planilha',
        )

        ParcelaEmprestimo.objects.create(
            emprestimo=emprestimo,
            numero=numero_parcela,
            valor=valor_decimal,
            vencimento=data_compra,
            pago=False,
        )

    def _normalize_date(self, value):
        if isinstance(value, datetime.datetime):
            return value.date()
        if isinstance(value, datetime.date):
            return value
        if isinstance(value, str):
            for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
                try:
                    return datetime.datetime.strptime(value.strip(), fmt).date()
                except Exception:
                    continue
        return None

    def _normalize_text(self, value):
        if value is None:
            return ''
        text = str(value)
        normalized = unicodedata.normalize('NFKD', text)
        return ''.join(c for c in normalized if not unicodedata.combining(c)).strip().upper()

    def _find_date_in_row(self, row):
        for candidate in row:
            if isinstance(candidate, datetime.datetime):
                return candidate.date()
            if isinstance(candidate, datetime.date):
                return candidate
            if isinstance(candidate, str):
                normalized = self._normalize_date(candidate)
                if normalized is not None:
                    return normalized

        # some rows may not contain a direct date; try the previous row or first non-empty cell
        for candidate in row:
            if isinstance(candidate, (int, float)):
                continue
            if isinstance(candidate, str):
                normalized = candidate.strip()
                if normalized:
                    try:
                        return datetime.datetime.strptime(normalized, '%d/%m/%Y').date()
                    except Exception:
                        continue
        return None

    def _get_or_create_categoria(self, nome, tipo):
        categoria, _ = Categoria.objects.get_or_create(nome=nome, tipo=tipo)
        return categoria

    def _get_or_create_conta_from_row(self, row):
        conta_nome = None
        banco_names = ['Mercado Pago', 'NuBank', 'Inter', 'Santander', 'Carteira']

        for candidate in row:
            if isinstance(candidate, str) and candidate in self.contas:
                conta_nome = candidate
                break

        if not conta_nome:
            for candidate in row:
                if isinstance(candidate, str) and candidate in banco_names:
                    conta_nome = candidate
                    break

        if not conta_nome:
            conta_nome = 'Carteira'

        return self.contas.get(conta_nome) or self.contas['Carteira']

    def _create_despesa(self, descricao, valor, data, categoria, conta, pago):
        Despesa.objects.create(
            descricao=str(descricao).strip(),
            valor=Decimal(str(valor)),
            data=data,
            categoria=categoria,
            conta=conta,
            pago=pago,
        )

    def _create_receita(self, descricao, valor, data, categoria, conta, recebido):
        Receita.objects.create(
            descricao=str(descricao).strip(),
            valor=Decimal(str(valor)),
            data=data,
            categoria=categoria,
            conta=conta,
            recebido=recebido,
        )
